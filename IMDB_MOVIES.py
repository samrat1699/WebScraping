from bs4 import BeautifulSoup
import requests
import openpyxl

# Creating Excel sheet 
excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = "Top Rated Movies"
print(excel.sheetnames)
sheet.append(["Movie Rank", "Movie Name", "Year of Release", "IMDB Rating"])


# Using try and except method to help you not showing error
try:
    source = requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')
    #print(soup)
# All find the movie from tbody parents to tr child here using two method find and find_all
    movies = soup.find('tbody', class_ = 'lister-list').find_all('tr')
    #print (len(movies))

    for movie in movies:

        name = movie.find('td', class_ = 'titleColumn').a.text

        rank = movie.find('td', class_ = 'titleColumn').get_text(strip = True).split('.')[0]

        year = movie.find('td', class_ = 'titleColumn').span.text.strip('()')

        rating = movie.find('td', class_ = 'ratingColumn imdbRating').strong.text

        print(rank, name, year, rating)
        sheet.append([rank, name, year, rating])

        

except Exception as e:
    print(e) 
# Save the Excel Sheet
excel.save('IMDB Movies Rating.xlsx')